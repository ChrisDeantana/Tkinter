import tkinter as tk
import PyPDF2
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfile
import os
from asyncio.windows_events import NULL
import openpyxl  # row and col starts from 1 - read
import xlsxwriter  # row and col starts from 0 - write
import random


def insert_input(difficulty, category, type, num, z, sheet_obj, list):
    row = sheet_obj.max_row - 2  # 30
    col = sheet_obj.max_column
    count = 0
    page_content = "Notice \n"
    list_save_id = []
    for i in range(0, row):
        for j in range(0, col):
            if (list[i][0] == difficulty) and (list[i][1] == category) and (list[i][2] == type):
                # print (list[i][3])
                list_save_id.append(list[i][3])
                count = count + 1
                break
    if len(list_save_id) < num:
        page_content = page_content + \
            f'There are only {str(len(list_save_id))} Questions that match your specification \n'
        # print("There are only " + str(len(list_save_id)) +
        #       " Questions that match your specification")
        # exit()
    # create new list with selected items
    new_list = [[0 for x in range(7)] for y in range(count)]

    cnt_x = 0
    cnt_y = 0
    for value in list_save_id:
        for i in range(0, row):
            if (list[i][3] == value):
                for j in range(4, col):
                    new_list[cnt_x][cnt_y] = list[i][j]
                    cnt_y = cnt_y + 1
                    if cnt_y == 7:
                        cnt_x = cnt_x + 1
                        cnt_y = 0
                        break

    # take random questions from new_list
    take_count = random.sample(list_save_id, num)

    take_list = [[0 for x in range(7)] for y in range(len(take_count))]

    cnt_x = 0
    cnt_y = 0
    for value in take_count:
        for i in range(0, row):
            if (list[i][3] == value):
                take_list[cnt_x][cnt_y] = list[i][4]
                cnt_y = cnt_y + 1
                random_list = [5, 6, 7, 8, 9]
                rnum = random.sample(random_list, len(random_list))
                for j in range(5):
                    take_list[cnt_x][cnt_y] = list[i][rnum[j]]
                    cnt_y = cnt_y + 1
                    if cnt_y == 6:
                        cnt_x = cnt_x + 1
                        cnt_y = 0
                        take_list[cnt_x-1][cnt_y-1] = list[i][10]
                        break

    desktop_path = os.path.join(os.path.join(
        os.environ['USERPROFILE']), 'Desktop')
    if (os.path.exists(f'{desktop_path}\Generated')):
        page_content += f'File has been succesfully generated in the {desktop_path}\Generated\n'
        print("no need to makedir")
    else:
        # page_content += f'A new folder called Generated has been created \n'
        os.makedirs(f'{desktop_path}\Generated')
        page_content += f'File has been succesfully generated in the {desktop_path}\Generated\n'
        print("need to makedir")

    workbook = xlsxwriter.Workbook(
        f'{desktop_path}/Generated/output' + str(z) + '.xlsx')
    worksheet = workbook.add_worksheet("My sheet")

    row = 0
    col = 0
    # Iterate over the data and write it out row by row.
    for val1, val2, val3, val4, val5, val6, val7 in (take_list):
        # print("output" + str(z) + '.xlsx')
        # print(f"[{val1}, {val2}, {val3}, {val4}, {val5}, {val6}, {val7}]")
        col = 0
        worksheet.write(row, col, val1)
        if val2 != None:
            col += 1
            worksheet.write(row, col, val2)
        if val3 != None:
            col += 1
            worksheet.write(row, col, val3)
        if val4 != None:
            col += 1
            worksheet.write(row, col, val4)
        if val5 != None:
            col += 1
            worksheet.write(row, col, val5)
        if val6 != None:
            col += 1
            worksheet.write(row, col, val6)
        worksheet.write(row, 6, val7)
        row += 1

    workbook.close()

    # text box
    text_box = tk.Text(root, height=10, width=40, padx=15, pady=15)
    text_box.insert(1.0, page_content)
    text_box.tag_configure("center", justify="center")
    text_box.tag_add("center", 1.0, "end")
    text_box.grid(columnspan=2, column=1, row=8, pady=25)

# def open_file():
#     browse_text.set("loading...")
#     file = askopenfile(parent=root, mode='rb', title="Choose a file", filetypes=[
#                        ("Pdf file", "*.pdf")])
#     print(file.name)
#     if file:
#         read_pdf = PyPDF2.PdfReader(file)
#         page = read_pdf.pages[int(selected_option.get())-1]
#         page_content = page.extract_text()

#         # text box
#         text_box = tk.Text(root, height=10, width=50, padx=15, pady=15)
#         text_box.insert(1.0, page_content)
#         text_box.tag_configure("center", justify="center")
#         text_box.tag_add("center", 1.0, "end")
#         text_box.grid(column=1, row=4)

#         # open file for writing
#         with open("example.txt", "w", encoding='utf-8') as file:
#             # write to file
#             file.write(page_content)

#         browse_text.set("Browse")


def open_file():
    file = askopenfile(parent=root, mode='rb', title="Choose a file", filetypes=[
                       ("Xlsx file", "*.xlsx")])
    # Processing the Data an Generate Algorithm
    # read .xlsx file
    path = file
    wb_obj = openpyxl.load_workbook(path)
    wb_obj.active = wb_obj['Sheet5']
    sheet_obj = wb_obj.active

    # col and row - the outliers
    row = sheet_obj.max_row - 2
    col = sheet_obj.max_column
    # create list with row x col
    list = [[0 for x in range(col)] for y in range(row)]

    # insert data to list
    for i in range(3, row + 3):
        for j in range(1, col + 1):
            list[i-3][j-1] = sheet_obj.cell(row=i, column=j).value
            # print(sheet_obj.cell(row = i, column = j).value)

    difficultyparam = selecteddiff_option.get()
    categoryparam = selectedcat_option.get()
    typeparam = selectedtype_option.get()
    N_of_Questionsparam = N_of_Questions_text_box.get("1.0", tk.END)
    NofQ = N_of_Questionsparam.strip()
    NofQ = int(NofQ)
    N_of_Sheetsparam = N_of_Sheets_text_box.get("1.0", tk.END)
    NofS = N_of_Sheetsparam.strip()
    NofS = int(NofS)

    for z in range(NofS):
        insert_input(difficultyparam, categoryparam,
                     typeparam, NofQ, z, sheet_obj, list)

    # print(difficultyparam)
    # print(categoryparam)
    # print(typeparam)
    # print(NofQ)
    # print(NofS)
    # print("endof open_file")


# THE GUI APPLICATION
root = tk.Tk()
root.title("MultipleChoiceRandomizer")
root.iconbitmap('Application\MultipleChoiceRandomizer\logo.ico')
root.resizable(False, False)

canvas = tk.Canvas(root, width=800, height=300)
canvas.grid(columnspan=4, rowspan=9)

# logo
logo = Image.open('Application\MultipleChoiceRandomizer\logo.png')
logo = ImageTk.PhotoImage(logo)
logo_label = tk.Label(image=logo)
# logo_label.image = logo
logo_label.grid(columnspan=2, column=1, row=0)

# difficulty
difficulty = tk.Label(root, text="Difficulty", font="Raleway")
difficulty.grid(column=1, row=1, sticky="e", padx=10, pady=5)
# category
Category = tk.Label(root, text="Category", font="Raleway")
Category.grid(column=1, row=2, sticky="e", padx=10, pady=5)
# type
Type = tk.Label(root, text="Type", font="Raleway")
Type.grid(column=1, row=3, sticky="e", padx=10, pady=5)
# N of Questions
N_of_Questions = tk.Label(root, text="N of Questions", font="Raleway")
N_of_Questions.grid(column=1, row=4, sticky="e", padx=10, pady=5)
# N of Sheets
N_of_Sheets = tk.Label(root, text="N of Sheets", font="Raleway")
N_of_Sheets.grid(column=1, row=5, sticky="e", padx=10, pady=5)

# difficulty selection
difficulty_options = ["A", "B", "C"]
selecteddiff_option = tk.StringVar(root)
selecteddiff_option.set(difficulty_options[0])
option_menu = tk.OptionMenu(root, selecteddiff_option, *difficulty_options)
option_menu.grid(column=2, row=1, sticky="w", padx=10, pady=5)

# category selection
category_options = ["A", "B"]
selectedcat_option = tk.StringVar(root)
selectedcat_option.set(category_options[0])
option_menu = tk.OptionMenu(root, selectedcat_option, *category_options)
option_menu.grid(column=2, row=2, sticky="w", padx=10, pady=5)

# type selection
type_options = ["M", "MM"]
selectedtype_option = tk.StringVar(root)
selectedtype_option.set(type_options[0])
option_menu = tk.OptionMenu(root, selectedtype_option, *type_options)
option_menu.grid(column=2, row=3, sticky="w", padx=10, pady=5)

# N_of_Questions
N_of_Questions_text_box = tk.Text(root, height=1, width=5)
N_of_Questions_text_box.grid(column=2, row=4, sticky="w", padx=10, pady=5)

# N_of_Sheets
N_of_Sheets_text_box = tk.Text(root, height=1, width=5)
N_of_Sheets_text_box.grid(column=2, row=5, sticky="w", padx=10, pady=5)

# instructions
instructions = tk.Label(root, text="Select your data", font="Raleway")
instructions.grid(columnspan=2, column=1, row=6, pady=(15, 0))

# browse button
browse_text = tk.StringVar()
browse_btn = tk.Button(root, textvariable=browse_text, command=lambda: open_file(),
                       font="Raleway", bg="#8fc649", fg="white", height=2, width=15)
browse_text.set("Browse")
browse_btn.grid(columnspan=2, column=1, row=7, sticky="N")

canvas = tk.Canvas(root, width=600, height=50)
canvas.grid(columnspan=3)

root.mainloop()
